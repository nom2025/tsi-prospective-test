"""
グラフ作成スクリプト

このスクリプトはCSVファイルからグラフを作成します。
再利用可能な関数として実装されており、様々なCSVファイルに対応できます。

実行方法:
    python ソース/create_graphs.py
"""

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from pathlib import Path
from datetime import datetime
import os

# seabornはオプショナル（インストールされていない場合はmatplotlibのデフォルトスタイルを使用）
try:
    import seaborn as sns
    sns.set_style("whitegrid")
    USE_SEABORN = True
except ImportError:
    USE_SEABORN = False
    # matplotlibのデフォルトスタイル設定
    plt.style.use('default')
    plt.rcParams['axes.grid'] = True
    plt.rcParams['grid.alpha'] = 0.3

# 日本語フォント設定（Windows環境）
# 利用可能な日本語フォントを順に試す
japanese_fonts = ['MS Gothic', 'MS PGothic', 'Yu Gothic', 'Meiryo', 'メイリオ', 'MS ゴシック']
font_found = False
for font in japanese_fonts:
    try:
        plt.rcParams['font.family'] = font
        # テスト用の文字列でフォントが正しく設定されるか確認
        fig_test = plt.figure(figsize=(1, 1))
        ax_test = fig_test.add_subplot(111)
        ax_test.text(0.5, 0.5, 'テスト', fontfamily=font)
        plt.close(fig_test)
        font_found = True
        break
    except:
        continue

if not font_found:
    # 日本語フォントが見つからない場合はデフォルトフォントを使用（警告を表示）
    import warnings
    warnings.warn("日本語フォントが見つかりません。日本語が正しく表示されない可能性があります。")
    plt.rcParams['font.family'] = 'DejaVu Sans'

# マイナス記号の文字化けを防ぐ
plt.rcParams['axes.unicode_minus'] = False

# プロジェクトルートのパスを取得
PROJECT_ROOT = Path(__file__).parent.parent
DATA_DIR = PROJECT_ROOT / "データ分析結果出力"
GRAPH_DIR = PROJECT_ROOT / "グラフ"

# グラフディレクトリが存在しない場合は作成
GRAPH_DIR.mkdir(parents=True, exist_ok=True)


def create_timeseries_graph(
    csv_file: Path,
    date_column: str = 'date',
    value_column: str = 'total_score',
    title: str = None,
    ylabel: str = None,
    output_filename: str = None,
    figsize: tuple = (20, 6),
    earthquake_data_file: Path = None,
    min_magnitude: float = 7.0
) -> Path:
    """
    時系列グラフを作成する
    
    Parameters:
    -----------
    csv_file : Path
        CSVファイルのパス
    date_column : str
        日付カラム名（デフォルト: 'date'）
    value_column : str
        値のカラム名（デフォルト: 'total_score'）
    title : str
        グラフのタイトル（Noneの場合は自動生成）
    ylabel : str
        Y軸のラベル（Noneの場合はvalue_columnを使用）
    output_filename : str
        出力ファイル名（Noneの場合は自動生成）
    figsize : tuple
        グラフのサイズ（デフォルト: (20, 6)）
    earthquake_data_file : Path
        地震データファイルのパス（M7.5以上の地震をプロットする場合）
    min_magnitude : float
        プロットする最小マグニチュード（デフォルト: 7.5）
    
    Returns:
    --------
    Path
        出力されたグラフファイルのパス
    """
    # CSVファイルを読み込み
    df = pd.read_csv(csv_file, parse_dates=[date_column])
    
    # 地震データを読み込んでM7以上の地震をプロット
    earthquake_markers = None
    if earthquake_data_file and earthquake_data_file.exists():
        try:
            # 地震データを読み込み
            eq_df = pd.read_csv(earthquake_data_file)
            
            # 日付変換関数（calculate_scores.pyと同じ）
            def parse_japan_earthquake_date(date_str):
                if '月' in str(date_str) and '日' in str(date_str) and '年' not in str(date_str):
                    return pd.to_datetime(f'2010年{date_str}', format='%Y年%m月%d日')
                else:
                    return pd.to_datetime(date_str, format='%Y/%m/%d')
            
            eq_df['date'] = eq_df['date'].apply(parse_japan_earthquake_date)
            eq_df['magnitude'] = pd.to_numeric(eq_df['magnitude'], errors='coerce')

            # M7.5以上の地震を抽出
            earthquake_markers = eq_df[eq_df['magnitude'] >= min_magnitude].copy()

            # 海外の地震を除外（日本国内のみ表示）
            # 除外する地域キーワード
            exclude_keywords = [
                '台湾', 'カムチャツカ', 'カムチャッカ', '小笠原諸島西方沖',
                'フィリピン', '南米', 'チリ', 'ペルー', 'インドネシア',
                'パプアニューギニア', 'ソロモン', 'バヌアツ', 'トンガ',
                'ニュージーランド', 'アラスカ', 'メキシコ', '中国',
                '千島列島', '択捉島南東沖'
            ]
            # 除外パターンに一致しない地震のみ残す
            exclude_pattern = '|'.join(exclude_keywords)
            earthquake_markers = earthquake_markers[
                ~earthquake_markers['place'].str.contains(exclude_pattern, na=False)
            ]

            earthquake_markers = earthquake_markers.sort_values('date')
        except Exception as e:
            print(f"警告: 地震データの読み込みに失敗しました: {e}")
            earthquake_markers = None
    
    # タイトルが指定されていない場合は自動生成（日本語）
    if title is None:
        # カラム名から日本語タイトルを自動生成
        title_map = {
            'total_score': '総合スコア',
            'japan_score': '国内地震スコア',
            'world_score': '世界地震スコア',
            'depth_score': '深さスコア',
            'gnss_score': 'GNSSスコア',
            'score': 'スコア'
        }
        title_base = title_map.get(value_column, value_column.replace('_', ' ').title())
        title = f"{title_base} - 時系列"
    
    # Y軸ラベルが指定されていない場合は自動生成（日本語）
    if ylabel is None:
        label_map = {
            'total_score': '総合スコア',
            'japan_score': '国内地震スコア',
            'world_score': '世界地震スコア',
            'depth_score': '深さスコア',
            'gnss_score': 'GNSSスコア',
            'score': 'スコア'
        }
        ylabel = label_map.get(value_column, value_column.replace('_', ' ').title())
    
    # 出力ファイル名が指定されていない場合は自動生成
    if output_filename is None:
        output_filename = f"{value_column}_timeseries.png"
    
    # グラフを作成
    fig, ax = plt.subplots(figsize=figsize)
    ax.plot(df[date_column], df[value_column], linewidth=1.5, color='#2E86AB', label=ylabel)
    ax.set_title(title, fontsize=16, fontweight='bold', pad=20)
    ax.set_xlabel('日付', fontsize=12)
    ax.set_ylabel(ylabel, fontsize=12)
    ax.grid(True, alpha=0.3)
    
    # M7.5以上の地震をプロット
    if earthquake_markers is not None and len(earthquake_markers) > 0:
        # まずグラフを描画してY軸の範囲を確定
        # Y軸の範囲を取得（アノテーション位置の計算に使用）
        y_min, y_max = ax.get_ylim()
        y_range = y_max - y_min

        # 日付の範囲を取得（日付間隔の計算に使用）
        date_min = df[date_column].min()
        date_max = df[date_column].max()
        date_range_days = (date_max - date_min).days

        # アノテーション位置を記録（重なり検出用）
        annotation_positions = []

        # 吹き出しのY位置レベル（5段階に増やして重なりを回避）
        y_levels = 5

        # 各M7.5以上の地震の日付に対応するスコアを取得
        for i, (idx, row) in enumerate(earthquake_markers.iterrows()):
            eq_date = row['date']
            eq_magnitude = row['magnitude']
            eq_place = row['place']

            # 該当日付のスコアを取得（最も近い日付を探す）
            closest_date_idx = (df[date_column] - eq_date).abs().idxmin()
            closest_date = df.loc[closest_date_idx, date_column]
            score_value = df.loc[closest_date_idx, value_column]

            # マーカーをグラフの線の上に少しオフセットしてプロット
            marker_offset = max(y_range * 0.02, 0.3)
            marker_y = score_value + marker_offset

            # マーカーをプロット（サイズを小さく）
            ax.scatter(eq_date, marker_y, s=120, color='red', marker='*',
                      edgecolors='darkred', linewidths=1, zorder=5,
                      label='M7以上' if i == 0 else '')

            # アノテーション（コンパクトに：場所とマグニチュードのみ）
            # 場所名が長い場合は短縮
            place_short = eq_place[:8] + '...' if len(eq_place) > 10 else eq_place
            annotation_text = f"{place_short}\nM{eq_magnitude:.1f}"

            # 現在の軸の範囲を取得
            xlim = ax.get_xlim()
            ylim = ax.get_ylim()
            y_range_current = ylim[1] - ylim[0]

            # 時間的に近い地震を検出して、レベルを分散させる
            # 既存のアノテーションとの距離をチェック
            best_level = i % y_levels
            min_collision_count = float('inf')

            for test_level in range(y_levels):
                collision_count = 0
                test_y = ylim[1] - y_range_current * 0.08 - (y_range_current * 0.12 * test_level)

                for prev_pos in annotation_positions:
                    x_diff_days = abs((eq_date - prev_pos['date']).days)
                    y_diff = abs(test_y - prev_pos['y'])

                    # 180日以内かつY方向が近い場合は衝突とみなす
                    if x_diff_days < 180 and y_diff < y_range_current * 0.12:
                        collision_count += 1

                if collision_count < min_collision_count:
                    min_collision_count = collision_count
                    best_level = test_level

            # Y位置を決定（上から順に5段階）
            annotation_y = ylim[1] - y_range_current * 0.08 - (y_range_current * 0.12 * best_level)

            # X位置：マーカーの少し右に配置
            x_offset = date_range_days * 0.02
            annotation_x = eq_date + pd.Timedelta(days=x_offset)

            # グラフの右端を超える場合は左に配置
            x_margin_days = date_range_days * 0.08
            if annotation_x > date_max - pd.Timedelta(days=x_margin_days):
                annotation_x = eq_date - pd.Timedelta(days=x_offset)
                ha_align = 'right'
            else:
                ha_align = 'left'

            # アノテーション位置を記録
            annotation_positions.append({
                'date': eq_date,
                'x': annotation_x,
                'y': annotation_y
            })

            # アノテーション（コンパクトなスタイル）
            ax.annotate(annotation_text,
                       xy=(eq_date, marker_y),
                       xytext=(annotation_x, annotation_y),
                       textcoords='data',
                       fontsize=7,
                       bbox=dict(boxstyle='round,pad=0.3', facecolor='lightyellow',
                                alpha=0.9, edgecolor='orange', linewidth=1),
                       arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.15',
                                      color='orange', lw=1, alpha=0.8),
                       ha=ha_align,
                       va='center',
                       zorder=6)

        # Y軸の範囲を調整（アノテーション用に上部に余白を追加）
        current_ylim = ax.get_ylim()
        y_padding_top = y_range * 0.25  # 上部余白を増やす
        y_padding_bottom = y_range * 0.02
        ax.set_ylim(current_ylim[0] - y_padding_bottom, current_ylim[1] + y_padding_top)
    
    # 日付のフォーマット設定（3か月毎）
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
    ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
    plt.xticks(rotation=45)

    # 凡例を左下に配置（吹き出しと重ならないようにする）
    if earthquake_markers is not None and len(earthquake_markers) > 0:
        ax.legend(loc='lower left', fontsize=9,
                 framealpha=0.95, edgecolor='gray',
                 fancybox=True, shadow=True)

    # レイアウト調整
    plt.tight_layout()
    
    # ファイルに保存
    output_path = GRAPH_DIR / output_filename
    # 既存ファイルを削除してタイムスタンプを確実に更新
    if output_path.exists():
        try:
            output_path.unlink()
        except Exception as e:
            print(f"警告: 既存ファイルの削除に失敗しました: {e}")
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    return output_path


def create_multiple_timeseries_graph(
    csv_files: list,
    date_column: str = 'date',
    value_columns: list = None,
    labels: list = None,
    title: str = "複数時系列データの比較",
    output_filename: str = "comparison_timeseries.png",
    figsize: tuple = (24, 8)
) -> Path:
    """
    複数の時系列データを1つのグラフに表示する
    
    Parameters:
    -----------
    csv_files : list
        CSVファイルのパスのリスト
    date_column : str
        日付カラム名（デフォルト: 'date'）
    value_columns : list
        値のカラム名のリスト（Noneの場合は各CSVの最初の数値カラムを使用）
    labels : list
        凡例のラベル（Noneの場合はファイル名から自動生成）
    title : str
        グラフのタイトル
    output_filename : str
        出力ファイル名
    figsize : tuple
        グラフのサイズ
    
    Returns:
    --------
    Path
        出力されたグラフファイルのパス
    """
    fig, ax = plt.subplots(figsize=figsize)
    
    colors = ['#2E86AB', '#A23B72', '#F18F01', '#C73E1D', '#6A994E']
    
    for i, csv_file in enumerate(csv_files):
        df = pd.read_csv(csv_file, parse_dates=[date_column])
        
        # 値のカラムを決定
        if value_columns is None:
            # 数値カラムを自動検出（dateカラム以外）
            numeric_cols = df.select_dtypes(include=['float64', 'int64']).columns.tolist()
            if date_column in numeric_cols:
                numeric_cols.remove(date_column)
            value_col = numeric_cols[0] if numeric_cols else None
        else:
            value_col = value_columns[i] if i < len(value_columns) else None
        
        if value_col is None:
            continue
        
        # ラベルを決定（日本語対応）
        if labels is None:
            # ファイル名から日本語ラベルを自動生成
            label_map = {
                'japan_scores': '国内地震スコア',
                'world_scores': '世界地震スコア',
                'depth_scores': '深さスコア',
                'gnss_scores': 'GNSSスコア',
                'total': '総合スコア'
            }
            label = label_map.get(csv_file.stem, csv_file.stem.replace('_', ' ').title())
        else:
            label = labels[i] if i < len(labels) else csv_file.stem
        
        # プロット
        ax.plot(df[date_column], df[value_col], 
                linewidth=1.5, label=label, color=colors[i % len(colors)])
    
    ax.set_title(title, fontsize=16, fontweight='bold', pad=20)
    ax.set_xlabel('日付', fontsize=12)
    ax.set_ylabel('スコア', fontsize=12)
    ax.legend(loc='best', fontsize=10)
    ax.grid(True, alpha=0.3)
    
    # 日付のフォーマット設定（3か月毎）
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
    ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
    plt.xticks(rotation=45)
    
    # レイアウト調整
    plt.tight_layout()
    
    # ファイルに保存
    output_path = GRAPH_DIR / output_filename
    # 既存ファイルを削除してタイムスタンプを確実に更新
    if output_path.exists():
        try:
            output_path.unlink()
        except Exception as e:
            print(f"警告: 既存ファイルの削除に失敗しました: {e}")
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    return output_path


def create_histogram(
    csv_file: Path,
    value_column: str = 'total_score',
    bins: int = 50,
    title: str = None,
    xlabel: str = None,
    output_filename: str = None,
    figsize: tuple = (10, 6)
) -> Path:
    """
    ヒストグラムを作成する
    
    Parameters:
    -----------
    csv_file : Path
        CSVファイルのパス
    value_column : str
        値のカラム名
    bins : int
        ビンの数（デフォルト: 50）
    title : str
        グラフのタイトル（Noneの場合は自動生成）
    xlabel : str
        X軸のラベル（Noneの場合はvalue_columnを使用）
    output_filename : str
        出力ファイル名（Noneの場合は自動生成）
    figsize : tuple
        グラフのサイズ
    
    Returns:
    --------
    Path
        出力されたグラフファイルのパス
    """
    # CSVファイルを読み込み
    df = pd.read_csv(csv_file)
    
    # タイトルが指定されていない場合は自動生成（日本語）
    if title is None:
        title_map = {
            'total_score': '総合スコア',
            'japan_score': '国内地震スコア',
            'world_score': '世界地震スコア',
            'depth_score': '深さスコア',
            'gnss_score': 'GNSSスコア',
            'score': 'スコア'
        }
        title_base = title_map.get(value_column, value_column.replace('_', ' ').title())
        title = f"{title_base} - 分布"
    
    # X軸ラベルが指定されていない場合は自動生成（日本語）
    if xlabel is None:
        label_map = {
            'total_score': '総合スコア',
            'japan_score': '国内地震スコア',
            'world_score': '世界地震スコア',
            'depth_score': '深さスコア',
            'gnss_score': 'GNSSスコア',
            'score': 'スコア'
        }
        xlabel = label_map.get(value_column, value_column.replace('_', ' ').title())
    
    # 出力ファイル名が指定されていない場合は自動生成
    if output_filename is None:
        output_filename = f"{value_column}_histogram.png"
    
    # グラフを作成
    fig, ax = plt.subplots(figsize=figsize)
    ax.hist(df[value_column], bins=bins, color='#2E86AB', edgecolor='black', alpha=0.7)
    ax.set_title(title, fontsize=16, fontweight='bold', pad=20)
    ax.set_xlabel(xlabel, fontsize=12)
    ax.set_ylabel('頻度', fontsize=12)
    ax.grid(True, alpha=0.3, axis='y')
    
    # 統計情報を表示
    mean_val = df[value_column].mean()
    median_val = df[value_column].median()
    ax.axvline(mean_val, color='red', linestyle='--', linewidth=2, label=f'平均: {mean_val:.2f}')
    ax.axvline(median_val, color='green', linestyle='--', linewidth=2, label=f'中央値: {median_val:.2f}')
    ax.legend()
    
    # レイアウト調整
    plt.tight_layout()
    
    # ファイルに保存
    output_path = GRAPH_DIR / output_filename
    # 既存ファイルを削除してタイムスタンプを確実に更新
    if output_path.exists():
        try:
            output_path.unlink()
        except Exception as e:
            print(f"警告: 既存ファイルの削除に失敗しました: {e}")
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    return output_path


def main():
    """メイン処理 - total.csvからグラフを作成"""
    import sys

    # コマンドライン引数の処理
    include_gnss = '--no-gnss' not in sys.argv

    # ファイル名のサフィックス（GNSSあり/なしを区別）
    suffix = "_with_gnss" if include_gnss else "_no_gnss"

    print("=" * 60)
    print("グラフ作成スクリプト")
    print("=" * 60)
    print(f"実行日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"データフォルダ: {DATA_DIR}")
    print(f"グラフ出力先: {GRAPH_DIR}")
    if not include_gnss:
        print("モード: GNSSスコアなし（--no-gnss）")
    else:
        print("モード: GNSSスコアあり")
    print("-" * 60)

    # total.csvのパス（GNSSあり/なしで異なるファイル）
    total_csv = DATA_DIR / f"total{suffix}.csv"

    if not total_csv.exists():
        print(f"[NG] エラー: ファイルが見つかりません: {total_csv}")
        return

    # 1. total.csvの時系列グラフを作成（M7.5以上の地震をプロット）
    gnss_label = "GNSSあり" if include_gnss else "GNSSなし"
    print(f"\n[1/3] total{suffix}.csvの時系列グラフを作成中（M7以上の地震をプロット）...")
    try:
        # 2010-all.csvのパス
        earthquake_data_file = PROJECT_ROOT / "生データ" / "2010-all.csv"

        output_path = create_timeseries_graph(
            csv_file=total_csv,
            date_column='date',
            value_column='total_score',
            title=f'総合スコア - 時系列（M7以上の地震表示）【{gnss_label}】',
            ylabel='総合スコア',
            output_filename=f'total_score_timeseries{suffix}.png',
            earthquake_data_file=earthquake_data_file,
            min_magnitude=7.0
        )
        print(f"[OK] 出力完了: {output_path}")
    except Exception as e:
        print(f"[NG] エラー: {e}")
        import traceback
        traceback.print_exc()
    
    # 2. total.csvのヒストグラムを作成
    print(f"\n[2/3] total{suffix}.csvのヒストグラムを作成中...")

    # 3. Excel形式でグラフデータを出力（Excelで編集可能）
    print("\n[3/3] Excel形式でグラフデータを出力中...")
    try:
        # openpyxlがインストールされているか確認
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.chart import LineChart, Reference

            # Excelファイルを作成
            excel_file = GRAPH_DIR / f"total_score_data{suffix}.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = f"総合スコアデータ【{gnss_label}】"

            # CSVデータを読み込んでExcelに書き込み
            df_total = pd.read_csv(total_csv, parse_dates=['date'])
            df_total['date_str'] = df_total['date'].dt.strftime('%Y-%m-%d')

            # ヘッダーを書き込み
            ws['A1'] = '日付'
            ws['B1'] = '総合スコア'

            # データを書き込み
            for idx, row in df_total.iterrows():
                ws.cell(row=idx+2, column=1, value=row['date_str'])
                ws.cell(row=idx+2, column=2, value=row['total_score'])

            # グラフを作成
            chart = LineChart()
            chart.title = f"総合スコア - 時系列【{gnss_label}】"
            chart.style = 10
            chart.y_axis.title = '総合スコア'
            chart.x_axis.title = '日付'

            # データ範囲を指定
            data = Reference(ws, min_col=2, min_row=1, max_row=len(df_total)+1)
            chart.add_data(data, titles_from_data=True)

            # X軸のカテゴリを指定
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(df_total)+1)
            chart.set_categories(cats)

            # グラフをワークシートに追加
            ws.add_chart(chart, "D2")

            # Excelファイルを保存
            wb.save(excel_file)
            print(f"[OK] Excel形式で出力完了: {excel_file}")
            print(f"  凡例の位置をExcelで編集できます")

        except ImportError:
            print("[!] openpyxlがインストールされていません。Excel形式での出力をスキップします")
            print("  インストール方法: pip install openpyxl")
    except Exception as e:
        print(f"[NG] Excel形式での出力エラー: {e}")
        import traceback
        traceback.print_exc()
    try:
        output_path = create_histogram(
            csv_file=total_csv,
            value_column='total_score',
            bins=50,
            title=f'総合スコア - 分布【{gnss_label}】',
            xlabel='総合スコア',
            output_filename=f'total_score_histogram{suffix}.png'
        )
        print(f"[OK] 出力完了: {output_path}")
    except Exception as e:
        print(f"[NG] エラー: {e}")
    
    # 3. 全スコアの比較グラフを作成（オプション）
    print("\n[オプション] 全スコアの比較グラフを作成中...")
    try:
        csv_files = [
            DATA_DIR / "japan_scores.csv",
            DATA_DIR / "world_scores.csv",
            DATA_DIR / "depth_scores.csv",
            DATA_DIR / "gnss_scores.csv"
        ]

        # GNSSなしモードの場合、gnss_scores.csvは除外
        if not include_gnss:
            csv_files = csv_files[:3]  # gnss_scores.csvを除外

        # ファイルが存在するか確認
        existing_files = [f for f in csv_files if f.exists()]

        if len(existing_files) >= 2:
            value_columns = ['score', 'world_score', 'depth_score', 'gnss_score']
            labels = ['国内地震スコア', '世界地震スコア', '深さスコア', 'GNSSスコア']
            # 存在するファイルに対応するvalue_columnsとlabelsを抽出
            value_columns = [value_columns[i] for i, f in enumerate(csv_files) if f.exists()]
            labels = [labels[i] for i, f in enumerate(csv_files) if f.exists()]

            output_path = create_multiple_timeseries_graph(
                csv_files=existing_files,
                value_columns=[value_columns[i] for i in range(len(existing_files))],
                labels=[labels[i] for i in range(len(existing_files))],
                title=f'スコア比較 - 時系列【{gnss_label}】',
                output_filename=f'all_scores_comparison{suffix}.png'
            )
            print(f"[OK] 出力完了: {output_path}")
        else:
            print("  スキップ: 必要なCSVファイルが不足しています")
    except Exception as e:
        print(f"[NG] エラー: {e}")

    print("\n" + "=" * 60)
    print("グラフ作成が完了しました！")
    print("=" * 60)
    print(f"\n出力ファイル【{gnss_label}】:")
    print(f"  - {GRAPH_DIR / f'total_score_timeseries{suffix}.png'} (総合スコア時系列、M7以上の地震表示)")
    print(f"  - {GRAPH_DIR / f'total_score_histogram{suffix}.png'} (総合スコア分布)")
    if (GRAPH_DIR / f'all_scores_comparison{suffix}.png').exists():
        if include_gnss:
            print(f"  - {GRAPH_DIR / f'all_scores_comparison{suffix}.png'} (全スコア比較: 国内+世界+深さ+GNSS)")
        else:
            print(f"  - {GRAPH_DIR / f'all_scores_comparison{suffix}.png'} (全スコア比較: 国内+世界+深さ)")


if __name__ == "__main__":
    main()

